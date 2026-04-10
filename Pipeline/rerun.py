"""
rerun.py — Move 1 Rerun Engine
-------------------------------
Given a parent deal_version (which points at a workbook in Supabase Storage),
apply assumption overrides, recalc the workbook with LibreOffice, read back
the canonical results per the Multifamily Results Contract, upload the new
workbook to Storage, and insert a new deal_versions row.

For the MVP, the editable assumptions are the four highest-leverage levers:
    - acquisition_price       → Assumptions!C7
    - ltv                     → Assumptions!C34
    - interest_rate           → Assumptions!C36
    - exit_cap_rate           → Assumptions!C45

The read-back cells follow the Multifamily Results Contract v1.1 (see
workspace doc Multifamily-Results-Contract.md). Anything else can be added
later without changing the rerun loop.
"""

import logging
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import supabase_client
from recalc import recalc_workbook, RecalcError

log = logging.getLogger("ping-server.rerun")

# ── Cell map for the four MVP levers ────────────────────────────────────────
# Keys are the override names the API accepts; values are (sheet, cell).
LEVER_CELLS: dict[str, tuple[str, str]] = {
    "acquisition_price":  ("Assumptions", "C7"),
    "ltv":                ("Assumptions", "C34"),
    "interest_rate":      ("Assumptions", "C36"),
    "exit_cap_rate":      ("Assumptions", "C45"),
}

# ── Read-back cell map (Multifamily Results Contract v1.1) ──────────────────
# Returns sheet C7-C12 are the canonical summary block when present;
# Pro Forma D-column gives Year-1 / stabilized values; Assumptions has
# the derived loan/equity numbers.
READBACK_MAP: dict[str, tuple[str, str]] = {
    # ── from Returns sheet (preferred) ──
    "levered_irr":        ("Returns",     "C7"),
    "moic":               ("Returns",     "C8"),
    "avg_coc":            ("Returns",     "C9"),
    "equity_multiple":    ("Returns",     "C10"),
    "total_profit":       ("Returns",     "C11"),
    "hold_period":        ("Returns",     "C12"),
    # ── from Pro Forma sheet ──
    "noi_stabilized":     ("Pro Forma",   "D21"),
    "cfbt_year1":         ("Pro Forma",   "D28"),
    "dscr":               ("Pro Forma",   "D32"),
    "coc_year1":          ("Pro Forma",   "D33"),
    "cap_rate_going_in":  ("Pro Forma",   "D34"),
    # ── from Assumptions sheet ──
    "acquisition_price":  ("Assumptions", "C7"),
    "total_acq_cost":     ("Assumptions", "C11"),
    "loan_amount":        ("Assumptions", "C35"),
    "equity_required":    ("Assumptions", "C58"),
    "ltv":                ("Assumptions", "C34"),
    "interest_rate":      ("Assumptions", "C36"),
    "exit_cap_rate":      ("Assumptions", "C45"),
}


class RerunError(RuntimeError):
    pass


def _coerce_number(val: Any) -> Any:
    """
    The contract specifies that IFERROR-wrapped formulas may return the
    string 'n/a'. Pass strings through; otherwise try to coerce to float.
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        return val
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def apply_overrides(workbook_path: Path, overrides: dict[str, float]) -> dict[str, Any]:
    """
    Open the workbook with formulas preserved, write the override values
    into the lever cells, and save. Returns a dict of {field: applied_value}
    for the values that were actually written (so the caller can record
    exactly what changed in deal_versions.assumption_overrides).
    """
    wb = load_workbook(workbook_path, data_only=False, keep_vba=False)
    applied: dict[str, Any] = {}

    for field, value in overrides.items():
        if field not in LEVER_CELLS:
            log.warning(f"rerun: ignoring unknown override field '{field}'")
            continue
        if value is None:
            continue
        sheet_name, cell = LEVER_CELLS[field]
        if sheet_name not in wb.sheetnames:
            raise RerunError(f"Workbook is missing required sheet '{sheet_name}'")
        ws = wb[sheet_name]
        ws[cell] = float(value)
        applied[field] = float(value)
        log.info(f"rerun: set {sheet_name}!{cell} = {value} ({field})")

    wb.save(workbook_path)
    wb.close()
    return applied


def read_back_results(workbook_path: Path) -> dict[str, Any]:
    """
    Open the recalculated workbook with data_only=True and pull every cell
    in the READBACK_MAP. Missing cells return None; 'n/a' strings pass through.
    """
    wb = load_workbook(workbook_path, data_only=True, keep_vba=False)
    out: dict[str, Any] = {}
    for field, (sheet_name, cell) in READBACK_MAP.items():
        if sheet_name not in wb.sheetnames:
            out[field] = None
            continue
        ws = wb[sheet_name]
        out[field] = _coerce_number(ws[cell].value)
    wb.close()
    return out


def run_rerun(
    *,
    deal_id: str,
    search_id: str,
    parent_workbook_path: str,
    overrides: dict[str, float],
    label: str,
    parent_version_id: str | None,
    created_by: str | None,
) -> dict[str, Any]:
    """
    The full rerun loop. Returns the inserted deal_versions row.

    Steps:
      1. Insert a 'pending' deal_versions row to get the version_id (we use
         it in the Storage path so the file is content-addressable to the row).
      2. Download parent workbook from Storage to a temp file.
      3. Apply overrides via openpyxl.
      4. Recalc via LibreOffice.
      5. Read back results.
      6. Upload new workbook to Storage at versions/{version_id}.xlsx.
      7. Patch the deal_versions row with workbook_path, results, status=complete.
      8. Return the row.

    Any failure flips status='failed' with an error message and re-raises.
    """
    # 1. Insert pending row
    version_row = supabase_client.insert_deal_version({
        "deal_id":              deal_id,
        "parent_version_id":    parent_version_id,
        "label":                label or "Untitled",
        "assumption_overrides": overrides or {},
        "status":               "pending",
        "created_by":           created_by,
    })
    version_id = version_row["id"]
    log.info(f"rerun: created pending version {version_id} for deal {search_id}")

    try:
        # 2. Download parent workbook
        with tempfile.TemporaryDirectory(prefix="rerun-") as tmp:
            tmp_dir = Path(tmp)
            local_path = tmp_dir / f"{search_id}_v{version_id[:8]}.xlsx"
            supabase_client.download_deal_file(parent_workbook_path, local_path)
            log.info(f"rerun: downloaded parent workbook ({local_path.stat().st_size} bytes)")

            # 3. Apply overrides
            applied = apply_overrides(local_path, overrides or {})

            # 4. Recalc
            recalc_workbook(local_path)
            log.info(f"rerun: recalc complete")

            # 5. Read back
            results = read_back_results(local_path)
            log.info(f"rerun: read back {len(results)} fields | IRR={results.get('levered_irr')}")

            # 6. Upload new workbook to Storage
            storage_path = f"{search_id}/versions/{version_id}.xlsx"
            supabase_client.upload_deal_file_to_path(storage_path, local_path)
            log.info(f"rerun: uploaded to {storage_path}")

        # 7. Patch row
        updated = supabase_client.update_deal_version(version_id, {
            "workbook_path":        storage_path,
            "results":              results,
            "assumption_overrides": applied,
            "status":               "complete",
        })
        return updated

    except (RecalcError, RerunError, Exception) as e:
        log.error(f"rerun: FAILED for version {version_id}: {e}", exc_info=True)
        try:
            supabase_client.update_deal_version(version_id, {
                "status": "failed",
                "error":  str(e)[:500],
            })
        except Exception:
            pass
        raise
