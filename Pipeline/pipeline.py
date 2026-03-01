"""
Ping Underwriting Pipeline
--------------------------
1. Fetch NEW searches from Apps Script (Google Sheets)
2. Call RentCast API for rental comps per unit-mix combo
3. Populate Excel model (Raw Comps + Rent Assumptions + Inputs)
4. Generate investment summary
5. Email the .xlsx to the requester
6. Mark Sheets row as DONE
"""

import json
import math
import os
import shutil
import smtplib
import sys
from collections import defaultdict
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Config ────────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
CONFIG_PATH = SCRIPT_DIR / "config.json"

with open(CONFIG_PATH) as f:
    CFG = json.load(f)

GAS_URL        = CFG["gas_url"]
RC_KEY         = CFG["rentcast_api_key"]
RC_BASE        = CFG["rentcast_base_url"]
TEMPLATE_PATH  = (SCRIPT_DIR / CFG["excel_template"]).resolve()
OUTPUT_DIR     = (SCRIPT_DIR / CFG["output_dir"]).resolve()
EMAIL_CFG      = CFG["email"]

# ── Helpers ───────────────────────────────────────────────────────────────────

def haversine_km(lat1, lon1, lat2, lon2):
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    return R * 2 * math.asin(math.sqrt(a))


def gas_get(action, **params):
    r = requests.get(GAS_URL, params={"action": action, **params}, timeout=30)
    r.raise_for_status()
    return r.json()


def rentcast_comps(lat, lng, radius, beds, baths, status, limit):
    params = {
        "latitude":  lat,
        "longitude": lng,
        "radius":    radius,
        "bedrooms":  beds,
        "bathrooms": baths,
        "status":    status,
        "limit":     min(int(limit), 500),
    }
    headers = {"X-Api-Key": RC_KEY}
    r = requests.get(f"{RC_BASE}/listings/rental", headers=headers, params=params, timeout=30)
    r.raise_for_status()
    return r.json()

# ── RentCast → Raw Comps rows ─────────────────────────────────────────────────

def build_comp_rows(listings, subject_lat, subject_lng, purchase_price, cost, filter_beds, filter_baths, unit_type):
    rows = []
    for listing in listings:
        comp_lat = listing.get("latitude") or 0
        comp_lng = listing.get("longitude") or 0
        dist_km  = haversine_km(subject_lat, subject_lng, comp_lat, comp_lng)
        dist_m   = dist_km * 1000
        rows.append({
            "purchase_price":  purchase_price,
            "improvements":    cost,
            "distance_m":      round(dist_m, 1),
            "distance_km":     round(dist_km, 3),
            "filter_beds":     filter_beds,
            "filter_baths":    filter_baths,
            "type":            unit_type,
            "formattedAddress": listing.get("formattedAddress", ""),
            "price":           listing.get("price", ""),
            "bedrooms":        listing.get("bedrooms", ""),
            "bathrooms":       listing.get("bathrooms", ""),
            "squareFootage":   listing.get("squareFootage", ""),
            "propertyType":    listing.get("propertyType", ""),
            "listing_status":  listing.get("status", ""),
            "daysOnMarket":    listing.get("daysOnMarket", ""),
            "latitude":        comp_lat,
            "longitude":       comp_lng,
            "url":             listing.get("listingUrl", ""),
            "id":              listing.get("id", ""),
        })
    rows.sort(key=lambda r: r["distance_m"])
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    return rows

# ── Excel Population ──────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="0F172A")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
DATA_FONT    = Font(name="Arial", size=9)
ALT_FILL     = PatternFill("solid", fgColor="F8FAFC")
BLUE_FONT    = Font(name="Arial", size=9, color="0000FF")   # hardcoded inputs
GREEN_FONT   = Font(name="Arial", size=9, color="008000")   # cross-sheet links

def set_cell(ws, row, col, value, font=None, fill=None, number_format=None, alignment=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:          cell.font          = font
    if fill:          cell.fill          = fill
    if number_format: cell.number_format = number_format
    if alignment:     cell.alignment     = alignment
    return cell


def populate_raw_comps(ws, all_rows, search_meta):
    """Overwrite Raw Comps data rows (keep header rows 1-3)."""
    # Clear existing data rows
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    col_headers = [
        "#", "Purchase Price", "Improvements", "rank",
        "distance_m", "distance_km", "filter_beds", "filter_baths", "type",
        "formattedAddress", "price", "bedrooms", "bathrooms", "squareFootage",
        "propertyType", "listing_status", "daysOnMarket",
        "latitude", "longitude", "url", "id"
    ]

    # Ensure header row 3 is correct (col B = col 2)
    for j, h in enumerate(col_headers):
        ws.cell(row=3, column=j+2, value=h).font = HEADER_FONT

    for i, row in enumerate(all_rows):
        r = i + 4
        fill = ALT_FILL if i % 2 == 0 else None
        vals = [
            i+1,
            row["purchase_price"],
            row["improvements"],
            row["rank"],
            row["distance_m"],
            row["distance_km"],
            row["filter_beds"],
            row["filter_baths"],
            row["type"],
            row["formattedAddress"],
            row["price"],
            row["bedrooms"],
            row["bathrooms"],
            row["squareFootage"],
            row["propertyType"],
            row["listing_status"],
            row["daysOnMarket"],
            row["latitude"],
            row["longitude"],
            row["url"],
            row["id"],
        ]
        for j, v in enumerate(vals):
            cell = ws.cell(row=r, column=j+2, value=v)
            cell.font = DATA_FONT
            if fill:
                cell.fill = fill


def populate_rent_assumptions(ws, search_meta, combos):
    """Update search metadata block (rows 2-4) in Rent Assumptions."""
    # Row 2: Search ID, lat
    ws.cell(row=2, column=2, value=search_meta["searchId"]).font = BLUE_FONT
    ws.cell(row=2, column=4, value=search_meta["lat"]).font      = BLUE_FONT
    ws.cell(row=2, column=6, value=search_meta["radius"]).font   = BLUE_FONT
    # Row 3: email, lng, maxComps
    ws.cell(row=3, column=2, value=search_meta["email"]).font    = BLUE_FONT
    ws.cell(row=3, column=4, value=search_meta["lng"]).font      = BLUE_FONT
    ws.cell(row=3, column=6, value=search_meta["maxComps"]).font = BLUE_FONT
    # Row 4: status, combo count, combo labels
    combo_label = " | ".join(f"{c['type']} {c['beds']}bd/{c['baths']}ba" for c in combos)
    ws.cell(row=4, column=2, value=search_meta["status"]).font       = BLUE_FONT
    ws.cell(row=4, column=4, value=len(combos)).font                 = BLUE_FONT
    ws.cell(row=4, column=6, value=combo_label).font                 = BLUE_FONT


def populate_inputs(ws, search_meta):
    """Update Inputs sheet with subject property data from the search."""
    # Walk rows to find labeled cells and update values
    for row in ws.iter_rows():
        for cell in row:
            v = str(cell.value or "").strip()
            if v == "Property Name":
                ws.cell(row=cell.row, column=cell.column+1, value=search_meta["address"]).font = BLUE_FONT
            elif v == "Acquisition Price ($)":
                price = search_meta.get("price")
                if price:
                    ws.cell(row=cell.row, column=cell.column+1, value=float(price)).font = BLUE_FONT


def aggregate_rent_assumptions(all_rows):
    """Group comp rows by combo type and compute avg rent + avg sqft."""
    groups = defaultdict(list)
    for r in all_rows:
        key = (r["type"], r["filter_beds"], r["filter_baths"])
        groups[key].append(r)

    summary = []
    for (unit_type, beds, baths), rows in groups.items():
        prices = [r["price"] for r in rows if r["price"]]
        sqfts  = [r["squareFootage"] for r in rows if r["squareFootage"]]
        summary.append({
            "type":     unit_type,
            "beds":     beds,
            "baths":    baths,
            "count":    len(rows),
            "avg_rent": round(sum(prices) / len(prices), 2) if prices else 0,
            "avg_sqft": round(sum(sqfts)  / len(sqfts),  2) if sqfts  else 0,
        })
    return summary

# ── Investment Summary ────────────────────────────────────────────────────────

def generate_summary(search_meta, combos, comp_summary):
    lines = []
    lines.append(f"INVESTMENT SUMMARY — {search_meta['address']}")
    lines.append(f"Search ID: {search_meta['searchId']}  |  {datetime.now().strftime('%B %d, %Y')}")
    lines.append("")
    lines.append("SUBJECT PROPERTY")
    if search_meta.get("price"):
        lines.append(f"  Acquisition Price: ${float(search_meta['price']):,.0f}")
    if search_meta.get("cost"):
        lines.append(f"  Estimated Improvements: ${float(search_meta['cost']):,.0f}")
    if search_meta.get("sqft"):
        lines.append(f"  Building SF: {float(search_meta['sqft']):,.0f}")
    lines.append(f"  Total Units: {search_meta['totalUnits']}")
    lines.append("")
    lines.append("RENTAL COMP ANALYSIS")
    lines.append(f"  Search Radius: {search_meta['radius']} mi  |  Status Filter: {search_meta['status']}")
    lines.append("")
    for s in comp_summary:
        lines.append(f"  {s['type']} ({s['beds']}bd/{s['baths']}ba)")
        lines.append(f"    Comps Found:    {s['count']}")
        lines.append(f"    Avg Rent/Mo:    ${s['avg_rent']:,.0f}")
        if s["avg_sqft"]:
            lines.append(f"    Avg SF:         {s['avg_sqft']:,.0f}")
            if s["avg_rent"] and s["avg_sqft"]:
                lines.append(f"    Rent / SF:      ${s['avg_rent']/s['avg_sqft']:.2f}")
        lines.append("")
    lines.append("The attached Excel model contains the full pro forma, debt schedule,")
    lines.append("returns analysis, and pricing scenarios populated with live comp data.")
    lines.append("")
    lines.append("— Ping Underwriting Engine")
    return "\n".join(lines)

# ── Email ─────────────────────────────────────────────────────────────────────

def send_email(to_addr, subject, body, attachment_path):
    if not EMAIL_CFG.get("enabled"):
        print("  [Email] Disabled in config — skipping send.")
        return

    msg = MIMEMultipart()
    msg["From"]    = EMAIL_CFG["sender"]
    msg["To"]      = to_addr
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    with open(attachment_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{Path(attachment_path).name}"')
    msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(EMAIL_CFG["sender"], EMAIL_CFG["app_password"])
        server.sendmail(EMAIL_CFG["sender"], to_addr, msg.as_string())
    print(f"  [Email] Sent to {to_addr}")

# ── Pipeline ──────────────────────────────────────────────────────────────────

def run_pipeline():
    print(f"\n{'='*60}")
    print(f"  Ping Pipeline — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}")

    # 1. Fetch new searches from Sheets
    print("\n[1] Fetching NEW searches from Google Sheets...")
    if GAS_URL == "YOUR_APPS_SCRIPT_WEB_APP_URL":
        print("  ❌  gas_url not configured in config.json — stopping.")
        sys.exit(1)

    data = gas_get("getNewSearches")
    rows = data.get("rows", [])
    if not rows:
        print("  No NEW searches found. Nothing to process.")
        return

    print(f"  Found {len(rows)} NEW row(s).")

    # Group rows by searchId (each combo = 1 row)
    searches = defaultdict(list)
    for row in rows:
        searches[row["searchId"]].append(row)

    print(f"  Grouped into {len(searches)} unique search(es).")

    for search_id, combo_rows in searches.items():
        print(f"\n[Search] {search_id}")

        # Mark as PROCESSING
        gas_get("updateStatus", searchId=search_id, status="PROCESSING")

        try:
            # Build search meta from first row (shared across combos)
            base = combo_rows[0]
            search_meta = {
                "searchId":   search_id,
                "email":      base["email"],
                "address":    base["address"],
                "lat":        float(base["lat"]),
                "lng":        float(base["lng"]),
                "radius":     float(base["radius"]),
                "minComps":   int(base["minComps"]),
                "maxComps":   int(base["maxComps"]),
                "status":     base["status"],
                "totalUnits": base["totalUnits"],
                "price":      base["price"],
                "cost":       base["cost"],
                "sqft":       base["sqft"],
            }
            combos = [{"beds": r["beds"], "baths": r["baths"], "type": r["type"]} for r in combo_rows]

            # 2. Fetch RentCast comps for each combo
            print(f"  [2] Fetching RentCast comps for {len(combos)} combo(s)...")
            all_comp_rows = []
            for combo in combos:
                print(f"      → {combo['type']} {combo['beds']}bd/{combo['baths']}ba")
                listings = rentcast_comps(
                    lat=search_meta["lat"],
                    lng=search_meta["lng"],
                    radius=search_meta["radius"],
                    beds=combo["beds"],
                    baths=combo["baths"],
                    status=search_meta["status"],
                    limit=search_meta["maxComps"],
                )
                comp_rows_for_combo = build_comp_rows(
                    listings,
                    search_meta["lat"], search_meta["lng"],
                    search_meta["price"], search_meta["cost"],
                    combo["beds"], combo["baths"], combo["type"],
                )
                print(f"         {len(comp_rows_for_combo)} comps returned.")
                all_comp_rows.extend(comp_rows_for_combo)

            # 3. Copy template and populate Excel
            print(f"  [3] Populating Excel model...")
            OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            ts         = datetime.now().strftime("%Y%m%d-%H%M")
            output_fn  = OUTPUT_DIR / f"Ping_{search_id}_{ts}.xlsx"
            shutil.copy(TEMPLATE_PATH, output_fn)

            wb = load_workbook(output_fn)
            populate_raw_comps(wb["Raw Comps"], all_comp_rows, search_meta)
            populate_rent_assumptions(wb["Rent Assumptions"], search_meta, combos)
            populate_inputs(wb["Inputs"], search_meta)
            wb.save(output_fn)
            print(f"         Saved: {output_fn.name}")

            # 4. Generate summary
            comp_summary = aggregate_rent_assumptions(all_comp_rows)
            summary_text = generate_summary(search_meta, combos, comp_summary)
            print(f"  [4] Investment summary generated.")

            # 5. Email
            print(f"  [5] Sending email to {search_meta['email']}...")
            subject = f"{EMAIL_CFG['subject_prefix']} — {base['address']}"
            send_email(search_meta["email"], subject, summary_text, output_fn)

            # 6. Mark DONE
            gas_get("updateStatus", searchId=search_id, status="DONE")
            print(f"  ✅  {search_id} complete.")

        except Exception as e:
            print(f"  ❌  Error processing {search_id}: {e}")
            gas_get("updateStatus", searchId=search_id, status="ERROR")

    print(f"\n{'='*60}")
    print("  Pipeline run complete.")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    run_pipeline()
