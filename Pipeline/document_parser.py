"""
document_parser.py — Parse commercial real estate documents and extract structured data.

Extends the lease parser to support three document types:
  1. Lease agreements         → 26 fields (existing)
  2. Offering Memorandums     → 32 fields (financial + property)
  3. Purchase & Sale contracts → 28 fields (deal terms + legal)

Accepts file bytes + filename + document_type, detects file format,
extracts text, and sends to OpenAI to extract type-specific fields as JSON.

Text-extraction strategy for PDFs (in order):
  1. pymupdf (fitz) — fast native text extraction; handles most modern PDFs.
  2. pytesseract OCR — for scanned / image-only PDFs where (1) yields < 100 chars.
"""

import io
import json
import logging
import os
from datetime import datetime

from openai import OpenAI

log = logging.getLogger("ping-server")


# ═══════════════════════════════════════════════════════════════════════════════
# FIELD SCHEMAS — one list per document type
# ═══════════════════════════════════════════════════════════════════════════════

LEASE_FIELDS = [
    "tenant_name",
    "tenant_entity_type",
    "property_address",
    "unit_number",
    "asset_class",
    "lease_start_date",
    "lease_end_date",
    "lease_term_months",
    "base_rent_monthly",
    "base_rent_annual",
    "rent_escalation_type",
    "rent_escalation_value",
    "free_rent_months",
    "security_deposit",
    "expense_structure",
    "tenant_responsible_expenses",
    "landlord_responsible_expenses",
    "tenant_improvement_allowance",
    "renewal_options",
    "termination_option",
    "termination_notice_months",
    "guarantor_name",
    "commencement_conditions",
    "notes",
]

OM_FIELDS = [
    # Property overview
    "property_name",
    "property_address",
    "asset_class",
    "property_type",
    "year_built",
    "year_renovated",
    "lot_size_acres",
    "building_sf",
    "total_units",
    "unit_mix",
    "occupancy_rate",
    "amenities",
    # Financial summary
    "asking_price",
    "price_per_unit",
    "price_per_sf",
    "cap_rate",
    "noi",
    "effective_gross_income",
    "operating_expenses",
    "expense_ratio",
    "gross_rent_multiplier",
    # Rent roll / income
    "average_rent_per_unit",
    "market_rent_per_unit",
    "rent_growth_potential",
    "other_income",
    "vacancy_loss",
    # Debt / returns
    "proposed_financing",
    "loan_to_value",
    "debt_service",
    "cash_on_cash_return",
    "projected_irr",
    # Context
    "seller_broker",
    "notes",
]

PSA_FIELDS = [
    # Parties
    "buyer_name",
    "buyer_entity_type",
    "seller_name",
    "seller_entity_type",
    # Property
    "property_address",
    "legal_description",
    "asset_class",
    "property_type",
    # Deal terms
    "purchase_price",
    "earnest_money_deposit",
    "additional_deposit",
    "deposit_escrow_agent",
    "closing_date",
    "due_diligence_period_days",
    "due_diligence_expiration",
    "financing_contingency",
    "financing_type",
    "loan_amount",
    "inspection_contingency",
    # Legal provisions
    "title_company",
    "closing_costs_allocation",
    "prorations",
    "representations_warranties",
    "default_remedies_buyer",
    "default_remedies_seller",
    "assignment_rights",
    "governing_law",
    "notes",
]


# ═══════════════════════════════════════════════════════════════════════════════
# EXTRACTION PROMPTS — one per document type
# ═══════════════════════════════════════════════════════════════════════════════

_PROMPT_TEMPLATE = (
    "You are a commercial real estate {role}. "
    "Extract the following fields from the provided {doc_label} text. "
    'Return a JSON object with these exact field names as keys. If a field is not found in the text, use "Not found" as the value. '
    "For numeric fields (prices, rates, percentages, counts), return the raw number without formatting — "
    "for example 1500000 not \"$1,500,000\" and 0.065 not \"6.5%\". "
    'Include a "confidence" field with an array of field names you are highly confident about. '
    'The "notes" field should contain any important caveats or ambiguities.\n\n'
    "{extra_instructions}"
    "Fields to extract:\n"
)

LEASE_EXTRACTION_PROMPT = (
    _PROMPT_TEMPLATE.format(
        role="lease analyst",
        doc_label="lease document",
        extra_instructions="",
    )
    + "\n".join(f"- {f}" for f in LEASE_FIELDS)
)

OM_EXTRACTION_PROMPT = (
    _PROMPT_TEMPLATE.format(
        role="investment analyst",
        doc_label="offering memorandum",
        extra_instructions=(
            'For "unit_mix", return a JSON array of objects like '
            '[{"type": "1BR/1BA", "count": 10, "avg_sf": 750, "avg_rent": 1200}]. '
            'For "proposed_financing", return a JSON object with keys like '
            '{"loan_amount": ..., "ltv": ..., "interest_rate": ..., "term_years": ..., "amortization_years": ...}. '
            'For "amenities", return a JSON array of strings. '
            "\n\n"
        ),
    )
    + "\n".join(f"- {f}" for f in OM_FIELDS)
)

PSA_EXTRACTION_PROMPT = (
    _PROMPT_TEMPLATE.format(
        role="transaction analyst",
        doc_label="purchase and sale agreement",
        extra_instructions=(
            'For "closing_costs_allocation", describe who pays which costs (e.g. "Seller pays transfer tax, Buyer pays title insurance"). '
            'For "prorations", describe what is prorated and how (e.g. "Taxes, insurance, and rents prorated as of closing"). '
            'For "representations_warranties", summarize key reps from both parties. '
            "\n\n"
        ),
    )
    + "\n".join(f"- {f}" for f in PSA_FIELDS)
)


# Map document_type → (fields, prompt)
DOC_TYPE_CONFIGS = {
    "lease": {
        "fields": LEASE_FIELDS,
        "prompt": LEASE_EXTRACTION_PROMPT,
        "label": "lease",
    },
    "om": {
        "fields": OM_FIELDS,
        "prompt": OM_EXTRACTION_PROMPT,
        "label": "offering memorandum",
    },
    "psa": {
        "fields": PSA_FIELDS,
        "prompt": PSA_EXTRACTION_PROMPT,
        "label": "purchase & sale agreement",
    },
}

SUPPORTED_DOC_TYPES = list(DOC_TYPE_CONFIGS.keys())


# ═══════════════════════════════════════════════════════════════════════════════
# TEXT EXTRACTION (reused from lease_parser.py — identical logic)
# ═══════════════════════════════════════════════════════════════════════════════

MIN_TEXT_CHARS = 100


def _extract_text_from_pdf(file_bytes: bytes) -> str:
    """Extract text from a PDF. Tries pymupdf first, falls back to OCR."""
    text = _extract_text_pymupdf(file_bytes)

    if len(text.strip()) >= MIN_TEXT_CHARS:
        return text.strip()

    log.info(
        "Native PDF text extraction yielded only %d chars — falling back to OCR",
        len(text.strip()),
    )
    ocr_text = _extract_text_ocr(file_bytes)
    if ocr_text and len(ocr_text.strip()) > len(text.strip()):
        return ocr_text.strip()

    return text.strip()


def _extract_text_pymupdf(file_bytes: bytes) -> str:
    """Extract text using pymupdf (fitz)."""
    try:
        import fitz

        pdf_doc = fitz.open(stream=file_bytes, filetype="pdf")
        text = ""
        for page in pdf_doc:
            text += page.get_text() + "\n"
        pdf_doc.close()
        return text
    except Exception as e:
        log.error(f"pymupdf text extraction failed: {e}")
        return ""


def _extract_text_ocr(file_bytes: bytes) -> str:
    """OCR a PDF using pdf2image + pytesseract."""
    try:
        from pdf2image import convert_from_bytes
        import pytesseract

        log.info("Running OCR on PDF (%d bytes)...", len(file_bytes))
        images = convert_from_bytes(file_bytes, dpi=300)
        text = ""
        for i, img in enumerate(images):
            page_text = pytesseract.image_to_string(img)
            text += page_text + "\n"
            log.info("OCR page %d: extracted %d chars", i + 1, len(page_text))
        return text
    except ImportError as e:
        log.warning("OCR dependencies not available (%s) — skipping OCR", e)
        return ""
    except Exception as e:
        log.error(f"OCR extraction failed: {e}")
        return ""


def _extract_text_from_csv(file_bytes: bytes) -> str:
    try:
        return file_bytes.decode("utf-8").strip()
    except Exception as e:
        log.error(f"Error extracting CSV text: {e}")
        raise


def _extract_text_from_xlsx(file_bytes: bytes) -> str:
    try:
        from openpyxl import load_workbook

        excel_file = io.BytesIO(file_bytes)
        workbook = load_workbook(excel_file)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"\n=== Sheet: {sheet_name} ===\n"
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        text += str(cell) + "\t"
                text += "\n"
        return text.strip()
    except Exception as e:
        log.error(f"Error extracting XLSX text: {e}")
        raise


def _extract_text_from_txt(file_bytes: bytes) -> str:
    try:
        return file_bytes.decode("utf-8").strip()
    except Exception as e:
        log.error(f"Error extracting TXT text: {e}")
        raise


def _extract_text(file_bytes: bytes, filename: str) -> str:
    """Detect file type and extract text accordingly."""
    filename_lower = filename.lower()
    if filename_lower.endswith(".pdf"):
        return _extract_text_from_pdf(file_bytes)
    elif filename_lower.endswith(".csv"):
        return _extract_text_from_csv(file_bytes)
    elif filename_lower.endswith(".xlsx"):
        return _extract_text_from_xlsx(file_bytes)
    elif filename_lower.endswith(".txt"):
        return _extract_text_from_txt(file_bytes)
    else:
        raise ValueError(f"Unsupported file type: {filename}")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINTS
# ═══════════════════════════════════════════════════════════════════════════════

def parse_document(file_bytes: bytes, filename: str, document_type: str = "lease") -> dict:
    """
    Parse a CRE document and extract type-specific fields using OpenAI.

    Args:
        file_bytes:    The file contents as bytes.
        filename:      The filename (used to detect file type).
        document_type: One of "lease", "om", "psa".

    Returns:
        {
            "document_type": str,
            "parsed": { ... extracted fields ... },
            "usage": {
                "prompt_tokens": int,
                "completion_tokens": int,
                "total_tokens": int,
                "estimated_cost": float,
            }
        }
    """
    if document_type not in DOC_TYPE_CONFIGS:
        raise ValueError(
            f"Unsupported document_type '{document_type}'. "
            f"Must be one of: {', '.join(SUPPORTED_DOC_TYPES)}"
        )

    config = DOC_TYPE_CONFIGS[document_type]

    # 1. Extract text
    log.info(f"Extracting text from {filename} (type: {config['label']})...")
    extracted_text = _extract_text(file_bytes, filename)
    log.info(f"Extracted {len(extracted_text)} characters from {filename}")

    if len(extracted_text.strip()) < 20:
        raise ValueError(
            f"Could not extract meaningful text from {filename}. "
            "The file may be empty, corrupted, or in an unsupported format."
        )

    # 2. Call OpenAI
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise ValueError("OPENAI_API_KEY environment variable not set")

    client = OpenAI(api_key=api_key)

    log.info(f"Sending extracted text to OpenAI gpt-4o-mini (type: {config['label']})...")
    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": config["prompt"]},
            {"role": "user", "content": extracted_text},
        ],
        response_format={"type": "json_object"},
        temperature=0,
    )

    # 3. Parse response
    response_text = response.choices[0].message.content
    parsed_data = json.loads(response_text)

    # gpt-4o-mini pricing: $0.15 / 1M input, $0.60 / 1M output
    prompt_cost = (response.usage.prompt_tokens / 1_000_000) * 0.15
    completion_cost = (response.usage.completion_tokens / 1_000_000) * 0.60
    total_cost = prompt_cost + completion_cost

    log.info(
        f"{config['label'].title()} parsing complete. "
        f"Tokens: {response.usage.total_tokens}, Cost: ${total_cost:.6f}"
    )

    return {
        "document_type": document_type,
        "parsed": parsed_data,
        "usage": {
            "prompt_tokens": response.usage.prompt_tokens,
            "completion_tokens": response.usage.completion_tokens,
            "total_tokens": response.usage.total_tokens,
            "estimated_cost": round(total_cost, 6),
        },
    }


def parse_lease(file_bytes: bytes, filename: str) -> dict:
    """Backward-compatible wrapper — delegates to parse_document(type='lease')."""
    result = parse_document(file_bytes, filename, document_type="lease")
    # Return in the old shape (no document_type key) for backward compat
    return {"parsed": result["parsed"], "usage": result["usage"]}
