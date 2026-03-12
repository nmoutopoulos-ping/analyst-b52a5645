"""
excel_writer.py — Ping Pipeline Excel Population
--------------------------------------------------
Writes live RentCast comp data and subject-property inputs into the
Multifamily Model Template.  Three sheet targets:

  Raw Comps   — every individual comp listing
  Assumptions — subject-property inputs + unit mix (drives all cash-flow formulas)
  Inputs      — display-only unit-mix summary panel
"""

from datetime import datetime

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from helpers import shorten_address

# ── Shared styles ──────────────────────────────────────────────────────────────
ALT_FILL  = PatternFill("solid", fgColor="F8FAFC")
BLUE_FONT = Font(name="Arial", size=9, color="0000FF")   # pipeline-written inputs


# ── Raw Comps ──────────────────────────────────────────────────────────────────

def populate_raw_comps(ws, all_rows: list, search_meta: dict) -> None:
    """
    Overwrite Raw Comps data rows (rows 5+). Rows 1-4 are untouched:
      Row 2 = sheet title
      Row 3 = subtitle / note
      Row 4 = column headers (baked into the template — not rewritten)

    Critical column positions — Assumptions AVERAGEIFS hard-reference these:
      Col I (9)  = price/rent      ← AVERAGEIFS range
      Col J (10) = filter_beds     ← AVERAGEIFS criteria
      Col K (11) = filter_baths    ← AVERAGEIFS criteria
      Col L (12) = squareFootage   ← AVERAGEIFS range (avg SF)

    Note: AVERAGEIFS formulas in Assumptions reference $I$4:$I$871 (starting at
    the header row). The header text in row 4 is safely ignored by AVERAGEIFS
    since its numeric criteria will never match a text cell.
    """
    from openpyxl.styles import Border, Side, Alignment as _Align
    from openpyxl.utils import get_column_letter as _gcl

    _thin   = Side(style="thin",   color="E2E8F0")
    _medium = Side(style="medium", color="94A3B8")
    _thin_b = Border(bottom=_thin)
    _med_b  = Border(bottom=_medium)
    _LEFT   = _Align(horizontal="left",   vertical="center")
    _CTR    = _Align(horizontal="center", vertical="center")
    _RIGHT  = _Align(horizontal="right",  vertical="center")
    _HDR_BG = PatternFill("solid", fgColor="1E3A5F")
    _HDR_F  = Font(bold=True, color="FFFFFF", name="Arial", size=8)
    _DAT_F  = Font(name="Arial", size=8, color="1F2937")

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # (label, col_width, data_align)
    # Column B = col 2 (col A is a narrow spacer)
    COL_SPEC = [
        ("#",              5,  _CTR),   # B  2
        ("Purchase Price", 16, _RIGHT), # C  3
        ("Improvements",   14, _RIGHT), # D  4
        ("Rank",           6,  _CTR),   # E  5
        ("Dist (m)",       9,  _RIGHT), # F  6
        ("Dist (km)",      9,  _RIGHT), # G  7
        ("Unit Type",      13, _LEFT),  # H  8
        ("Rent / Mo ★",   12, _RIGHT), # I  9  ← AVERAGEIFS range
        ("Beds ★",         7,  _CTR),   # J  10 ← AVERAGEIFS criteria
        ("Baths ★",        8,  _CTR),   # K  11 ← AVERAGEIFS criteria
        ("Sq Ft ★",        9,  _RIGHT), # L  12 ← AVERAGEIFS range
        ("Bedrooms",       10, _CTR),   # M  13
        ("Bathrooms",      10, _CTR),   # N  14
        ("Address",        34, _LEFT),  # O  15
        ("Property Type",  14, _LEFT),  # P  16
        ("Status",         10, _CTR),   # Q  17
        ("DOM",            7,  _CTR),   # R  18
        ("Latitude",       11, _RIGHT), # S  19
        ("Longitude",      11, _RIGHT), # T  20
        ("URL",            12, _LEFT),  # U  21
        ("Listing ID",     18, _LEFT),  # V  22
    ]
    COL_FMTS = {
        3: "$#,##0", 4: "$#,##0", 5: "0", 6: "0", 7: "0.00",
        9: "$#,##0", 10: "0",     11: "0.00", 12: "0",
        13: "0",     14: "0.00",  19: "0.00", 20: "0.00",
    }
    VALS = [
        lambda r: r["purchase_price"],
        lambda r: r["improvements"],
        lambda r: r["rank"],
        lambda r: r["distance_m"],
        lambda r: r["distance_km"],
        lambda r: r["type"],
        lambda r: r["price"],
        lambda r: r["filter_beds"],
        lambda r: r["filter_baths"],
        lambda r: r["squareFootage"],
        lambda r: r["bedrooms"],
        lambda r: r["bathrooms"],
        lambda r: r["formattedAddress"],
        lambda r: r["propertyType"],
        lambda r: r["listing_status"],
        lambda r: r["daysOnMarket"],
        lambda r: r["latitude"],
        lambda r: r["longitude"],
        lambda r: r["url"],
        lambda r: r["id"],
    ]

    # Set column widths; row 4 headers are preserved from the template (not rewritten).
    for j, (_, width, _) in enumerate(COL_SPEC):
        col = j + 2
        ws.column_dimensions[_gcl(col)].width = width

    ws.column_dimensions["A"].width = 0.5
    ws.freeze_panes = "B5"

    for i, row in enumerate(all_rows):
        r    = i + 5
        fill = ALT_FILL if i % 2 == 0 else None
        idx  = ws.cell(row=r, column=2, value=i + 1)
        idx.font = _DAT_F; idx.alignment = _CTR
        if fill:
            idx.fill = fill
        for j, fn in enumerate(VALS):
            col = j + 3
            _, _, align = COL_SPEC[j + 1]
            v  = fn(row)
            cl = ws.cell(row=r, column=col, value=v)
            cl.font      = _DAT_F
            cl.alignment = align
            cl.border    = _thin_b
            fmt = COL_FMTS.get(col)
            if fmt and v not in (None, ""):
                cl.number_format = fmt
            if fill and v is not None:
                cl.fill = fill



# ── Assumptions ───────────────────────────────────────────────────────────────

def populate_assumptions(ws, search_meta: dict, combos: list,
                         comp_summary: list,
                         commercial_spaces: list = None) -> None:
    """
    Write subject-property inputs to the Assumptions sheet.
    This sheet drives Pro Forma, Returns, and Pricing Scenarios via
    cross-sheet formulas.

    Header block (rows 3-4):
      C3 = Property Name (short address)
      C4 = "Search ID  |  Full address"

    Model inputs (drive all downstream formulas):
      C7  = Acquisition Price
      C10 = CapEx / Renovation Budget
      C12 = Total Units
      B16:D20 = unit-type labels + beds + baths (drive AVERAGEIFS comp lookup)
    """
    short = shorten_address(search_meta["address"])
    ws["C3"] = short; ws["C3"].font = BLUE_FONT

    search_id = search_meta.get("searchId", "—")
    address   = search_meta.get("address", "—")
    ws["C4"] = f"{search_id}  |  {address}"; ws["C4"].font = BLUE_FONT

    try:
        ws["C7"] = float(search_meta["price"]); ws["C7"].font = BLUE_FONT
    except (ValueError, TypeError):
        pass
    try:
        ws["C10"] = float(search_meta["cost"]); ws["C10"].font = BLUE_FONT
    except (ValueError, TypeError):
        pass
    try:
        ws["C12"] = int(float(search_meta["totalUnits"])); ws["C12"].font = BLUE_FONT
    except (ValueError, TypeError):
        pass

    seen, unique = set(), []
    for c in combos:
        key = (c["beds"], c["baths"])
        if key not in seen:
            seen.add(key); unique.append(c)

    for i in range(20 - 16 + 1):
        r = 16 + i
        if i < len(unique):
            c = unique[i]
            try:
                beds = int(float(c["beds"]))
                bf   = float(c["baths"])
                ba   = int(bf) if bf == int(bf) else bf
            except (ValueError, TypeError):
                beds, ba = c["beds"], c["baths"]
            ws.cell(row=r, column=2, value=f"{beds}BR/{ba}BA").font = BLUE_FONT
            ws.cell(row=r, column=3, value=float(c["beds"])).font   = BLUE_FONT
            ws.cell(row=r, column=4, value=float(c["baths"])).font  = BLUE_FONT
            # Write actual unit count to col E, overriding the template formula
            try:
                units_val = int(float(c.get("units", 0) or 0))
            except (ValueError, TypeError):
                units_val = 0
            ws.cell(row=r, column=5, value=units_val if units_val > 0 else None).font = BLUE_FONT
        else:
            ws.cell(row=r, column=2, value=None)
            ws.cell(row=r, column=3, value=None)
            ws.cell(row=r, column=4, value=None)
            ws.cell(row=r, column=5, value=None)

    # ── Commercial Spaces (rows 64+) ──────────────────────────────────────────
    if commercial_spaces:
        _NAV_BG = PatternFill("solid", fgColor="1E3A5F")
        _NAV_F  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        _GRY_BG = PatternFill("solid", fgColor="F1F5F9")
        _GRY_F  = Font(bold=True, color="475569", name="Arial", size=8)
        _DAT2   = Font(name="Arial", size=9, color="1F2937")

        SECT = 64
        CH   = 65
        D0   = 66   # first data row

        # Section header
        c = ws.cell(row=SECT, column=2, value="⑩ COMMERCIAL SPACES ($/SF/Yr — NNN)")
        c.font = _NAV_F; c.fill = _NAV_BG

        # Column headers
        for col, lbl in [(2, "Space Type"), (3, "SF"), (4, "$/SF/Yr"), (5, "Gross Rent/Yr")]:
            cell = ws.cell(row=CH, column=col, value=lbl)
            cell.font = _GRY_F; cell.fill = _GRY_BG

        total_sf = 0; total_gross = 0.0
        for i, space in enumerate(commercial_spaces[:5]):
            r = D0 + i
            try:
                sf    = int(float(space.get("sqft",     0) or 0))
                rpsf  = float(space.get("rentPerSF", 0) or 0)
                gross = sf * rpsf
            except (ValueError, TypeError):
                sf, rpsf, gross = 0, 0, 0.0
            total_sf += sf; total_gross += gross
            fill = ALT_FILL if i % 2 == 0 else None
            for col, val, fmt in [
                (2, space.get("type", ""), None),
                (3, sf or None,            "#,##0"),
                (4, rpsf or None,          "$#,##0.00"),
                (5, gross or None,         "$#,##0"),
            ]:
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = BLUE_FONT
                if fmt and val is not None: cell.number_format = fmt
                if fill: cell.fill = fill

        # Totals row
        tot_row = D0 + len(commercial_spaces)
        ws.cell(row=tot_row, column=2, value="TOTAL").font = Font(bold=True, name="Arial", size=9, color="1F2937")
        c_sf = ws.cell(row=tot_row, column=3, value=total_sf or None)
        c_sf.font = Font(bold=True, name="Arial", size=9, color="1F2937")
        c_sf.number_format = "#,##0"
        c_gr = ws.cell(row=tot_row, column=5, value=round(total_gross, 0) or None)
        c_gr.font = Font(bold=True, name="Arial", size=9, color="059669" if total_gross > 0 else "1F2937")
        c_gr.number_format = "$#,##0"


# ── Inputs ─────────────────────────────────────────────────────────────────────

def populate_inputs(ws, search_meta: dict, combos: list,
                    comp_summary: list,
                    commercial_spaces: list = None) -> None:
    """
    Write display data to the Inputs sheet (informational only — does not
    drive Pro Forma formulas; those reference Assumptions).

    Row 5:  B = Property Name
    Row 6:  B = Acquisition Price
    Rows 13-19: unit mix table (Type, Beds, Baths, #Units, AvgSF, AvgRent, Rev)
    Row 26: B = Renovation Cost / Unit
    """
    short = shorten_address(search_meta["address"])
    ws.cell(row=5, column=2, value=short).font = BLUE_FONT
    try:
        ws.cell(row=6, column=2, value=float(search_meta["price"])).font = BLUE_FONT
    except (ValueError, TypeError):
        pass
    try:
        total_units   = int(float(search_meta["totalUnits"]))
        cost_per_unit = float(search_meta["cost"]) / total_units if total_units else 0
        ws.cell(row=26, column=2, value=round(cost_per_unit, 0)).font = BLUE_FONT
    except (ValueError, TypeError):
        pass

    comp_lookup = {(str(s["beds"]), str(s["baths"])): s for s in comp_summary}

    seen, unique = set(), []
    for c in combos:
        key = (c["beds"], c["baths"])
        if key not in seen:
            seen.add(key); unique.append(c)

    for i in range(19 - 13 + 1):
        r = 13 + i
        if i < len(unique):
            c = unique[i]
            try:
                beds = int(float(c["beds"]))
                bf   = float(c["baths"])
                ba   = int(bf) if bf == int(bf) else bf
            except (ValueError, TypeError):
                beds, ba = c["beds"], c["baths"]
            cs = comp_lookup.get((str(c["beds"]), str(c["baths"])), {})
            # Use actual unit count from combo data
            try:
                n_units = int(float(c.get("units", 0) or 0)) or None
            except (ValueError, TypeError):
                n_units = None
            label   = f"{beds}BR/{ba}BA"
            ws.cell(row=r, column=1, value=label).font          = BLUE_FONT
            ws.cell(row=r, column=2, value=beds).font           = BLUE_FONT
            ws.cell(row=r, column=3, value=ba).font             = BLUE_FONT
            ws.cell(row=r, column=4, value=n_units).font        = BLUE_FONT
            ws.cell(row=r, column=5, value=cs.get("avg_sqft") or None).font = BLUE_FONT
            ws.cell(row=r, column=6, value=cs.get("avg_rent") or None).font = BLUE_FONT
            d = ws.cell(row=r, column=4).column_letter
            f = ws.cell(row=r, column=6).column_letter
            g = ws.cell(row=r, column=7).column_letter
            ws.cell(row=r, column=7, value=f"={d}{r}*{f}{r}").font = Font(name="Arial", size=9)
            ws.cell(row=r, column=8, value=f"={g}{r}*12").font     = Font(name="Arial", size=9)
        else:
            for col in range(1, 9):
                ws.cell(row=r, column=col).value = None

    # ── Commercial Leasing Section (rows 80+) ─────────────────────────────────
    if commercial_spaces:
        # Aggregate totals
        total_sf = 0; total_gross = 0.0
        for s in commercial_spaces:
            try:
                sf   = int(float(s.get("sqft",     0) or 0))
                rpsf = float(s.get("rentPerSF", 0) or 0)
                total_sf    += sf
                total_gross += sf * rpsf
            except (ValueError, TypeError):
                pass

        # ── Styles ──
        _NAV_BG = PatternFill("solid", fgColor="1E3A5F")
        _NAV_F  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        _HDR_BG = PatternFill("solid", fgColor="F1F5F9")
        _HDR_F  = Font(bold=True, color="475569", name="Arial", size=8)
        _TOT_BG = PatternFill("solid", fgColor="E2E8F0")
        _DAT3   = Font(name="Arial", size=9, color="1F2937")

        R_SECT = 81
        R_CH   = 82
        R_D0   = 83
        R_TOT  = R_D0 + len(commercial_spaces)

        # Section header
        c = ws.cell(row=R_SECT, column=1, value="⑩ COMMERCIAL SPACES")
        c.font = _NAV_F; c.fill = _NAV_BG

        # Column headers
        for col, lbl in [(1, "Space Type"), (2, "Total SF"),
                         (3, "$/SF/Yr"), (4, "Gross Annual Rent")]:
            cell = ws.cell(row=R_CH, column=col, value=lbl)
            cell.font = _HDR_F; cell.fill = _HDR_BG

        # Data rows
        for i, s in enumerate(commercial_spaces[:5]):
            r = R_D0 + i
            try:
                sf    = int(float(s.get("sqft",     0) or 0))
                rpsf  = float(s.get("rentPerSF", 0) or 0)
                gross = sf * rpsf
            except (ValueError, TypeError):
                sf, rpsf, gross = 0, 0, 0.0
            fill = ALT_FILL if i % 2 == 0 else None
            for col, val, fmt in [
                (1, s.get("type", ""),  None),
                (2, sf or None,         "#,##0"),
                (3, rpsf or None,       "$#,##0.00"),
                (4, gross or None,      "$#,##0"),
            ]:
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = _DAT3
                if fmt and val is not None: cell.number_format = fmt
                if fill: cell.fill = fill

        # Totals row
        for col, val, fmt, bold in [
            (1, "TOTAL",              None,     True),
            (2, total_sf or None,     "#,##0",  True),
            (3, None,                 None,     False),
            (4, total_gross or None,  "$#,##0", True),
        ]:
            cell = ws.cell(row=R_TOT, column=col, value=val)
            cell.font = Font(bold=bold, name="Arial", size=9, color="1F2937")
            cell.fill = _TOT_BG
            if fmt and val is not None: cell.number_format = fmt
