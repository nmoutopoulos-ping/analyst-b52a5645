"""
resend_email.py — Resend the email for an existing deal output.

Reads search metadata, comp summary, and commercial spaces directly from
the already-generated Excel model so no API calls are made. The pipeline,
template, and all other files are left completely untouched.

Usage:
    # Resend the most recently generated deal:
    python3 resend_email.py

    # Resend a specific search ID:
    python3 resend_email.py SRCH-20260304-001

    # Preview the email body without sending:
    python3 resend_email.py --preview
    python3 resend_email.py SRCH-20260304-001 --preview
"""

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from emailer import generate_email_body, send_email
from helpers import shorten_address


OUTPUT_DIR = Path(__file__).parent / "output"


# ── Locate the output folder ──────────────────────────────────────────────────

def find_job_dir(search_id: str | None) -> Path:
    """Return the job folder for the given search_id, or the most recent one."""
    folders = sorted(OUTPUT_DIR.iterdir(), key=lambda p: p.stat().st_mtime, reverse=True)
    folders = [f for f in folders if f.is_dir()]

    if not folders:
        raise FileNotFoundError(f"No output folders found in {OUTPUT_DIR}")

    if search_id:
        matches = [f for f in folders if search_id in f.name]
        if not matches:
            raise FileNotFoundError(f"No output folder found for search ID: {search_id}")
        return matches[0]

    return folders[0]


# ── Parse folder name for search_id + email ───────────────────────────────────

def parse_folder_meta(job_dir: Path) -> dict:
    """
    Folder name format: "SRCH-YYYYMMDD-NNN — Short Address — user@example.com"
    Extract search_id and email from the name — these aren't stored in the sheet.
    """
    parts = job_dir.name.split(" — ")
    search_id = parts[0].strip() if parts else ""
    email     = parts[-1].strip() if len(parts) >= 3 else ""
    return {"searchId": search_id, "email": email}


# ── Reconstruct search_meta from the Assumptions sheet ───────────────────────

def load_search_meta(wb, folder_meta: dict) -> dict:
    """
    Current Assumptions layout (written by excel_writer.populate_assumptions):
      C3  = short address
      C4  = "SRCH-YYYYMMDD-NNN  |  Full address"
      C7  = Acquisition Price
      C10 = CapEx / Renovation Budget
      C12 = Total Units
    Email and search_id come from the folder name (not stored in the sheet).
    """
    ws = wb["Assumptions"]

    # Address — C4 has "search_id  |  full address"; fall back to C3
    c4 = str(ws["C4"].value or "")
    if "|" in c4:
        address = c4.split("|", 1)[1].strip()
    else:
        address = str(ws["C3"].value or "").strip()

    price, cost, total_units = None, None, None
    try: price       = float(ws["C7"].value  or 0)
    except (TypeError, ValueError): pass
    try: cost        = float(ws["C10"].value or 0)
    except (TypeError, ValueError): pass
    try: total_units = int(float(ws["C12"].value or 0))
    except (TypeError, ValueError): pass

    # Radius / status — try to parse from the Raw Comps subtitle row (row 3)
    radius, status = 10.0, "Active"
    if "Raw Comps" in wb.sheetnames:
        subtitle = str(wb["Raw Comps"].cell(row=3, column=1).value or "")
        m = re.search(r"([\d.]+)\s*mi", subtitle)
        if m:
            try: radius = float(m.group(1))
            except ValueError: pass
        for st in ("Active", "Pending", "Inactive", "Off Market"):
            if st.lower() in subtitle.lower():
                status = st
                break

    return {
        "searchId":   folder_meta.get("searchId", ""),
        "email":      folder_meta.get("email", ""),
        "address":    address,
        "radius":     radius,
        "status":     status,
        "price":      str(price      or 0),
        "cost":       str(cost       or 0),
        "totalUnits": str(total_units or 0),
        "sqft":       "0",
    }


# ── Reconstruct comp_summary from Raw Comps ───────────────────────────────────

def load_comp_summary(wb) -> list:
    """
    Raw Comps data rows start at row 5. Critical columns (per excel_writer.py):
      Col I  (9)  = rent/mo
      Col J  (10) = filter_beds
      Col K  (11) = filter_baths
      Col L  (12) = squareFootage
    Group rows by (beds, baths) and compute averages.
    """
    ws = wb["Raw Comps"]
    buckets: dict = {}

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        rent  = row[8]   # col I
        beds  = row[9]   # col J
        baths = row[10]  # col K
        sqft  = row[11]  # col L

        if beds is None or baths is None or rent is None:
            continue
        try:
            beds  = int(float(beds))
            baths = float(baths)
            rent  = float(rent)
            sqft  = float(sqft) if sqft else None
        except (TypeError, ValueError):
            continue

        key = (beds, baths)
        if key not in buckets:
            buckets[key] = {"rents": [], "sqfts": [], "count": 0}
        buckets[key]["rents"].append(rent)
        buckets[key]["count"] += 1
        if sqft:
            buckets[key]["sqfts"].append(sqft)

    summary = []
    for (beds, baths), data in sorted(buckets.items(), key=lambda x: (-x[0][0], -x[0][1])):
        avg_rent = sum(data["rents"]) / len(data["rents"]) if data["rents"] else 0
        avg_sqft = sum(data["sqfts"]) / len(data["sqfts"]) if data["sqfts"] else None
        baths_str = str(int(baths) if baths == int(baths) else baths)
        summary.append({
            "beds":     str(beds),
            "baths":    baths_str,
            "count":    data["count"],
            "avg_rent": round(avg_rent, 2),
            "avg_sqft": round(avg_sqft, 1) if avg_sqft else None,
        })

    return summary


# ── Reconstruct commercial spaces from Assumptions rows 64+ ──────────────────

def load_commercial_spaces(wb) -> list:
    """
    Commercial section written by excel_writer.populate_assumptions.
    Scans col B of Assumptions for the section header, skips the col-header row,
    then reads data rows until a TOTAL row or empty col B is encountered.
    Layout (col B): header → "Space Type" → data rows → "TOTAL"
    Data cols: B=type, C=sf, D=rpsf, E=gross_annual
    """
    ws = wb["Assumptions"]
    spaces = []

    # Find the section header row by scanning col B up to row 100
    header_row = None
    for r in range(1, 100):
        val = ws.cell(row=r, column=2).value
        if val and "COMMERCIAL SPACES" in str(val).upper():
            header_row = r
            break

    if header_row is None:
        return []   # no commercial section in this model

    # Data starts 2 rows after the section header (skip col-header row)
    data_start = header_row + 2

    for row_idx in range(data_start, data_start + 6):   # max 5 spaces
        type_val = ws.cell(row=row_idx, column=2).value
        if type_val is None or str(type_val).strip().upper() == "TOTAL":
            break
        sf_val   = ws.cell(row=row_idx, column=3).value
        rpsf_val = ws.cell(row=row_idx, column=4).value
        try:
            spaces.append({
                "type":      str(type_val).strip(),
                "sqft":      str(int(float(sf_val   or 0))),
                "rentPerSF": str(float(rpsf_val or 0)),
            })
        except (TypeError, ValueError):
            continue

    return spaces


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    args    = sys.argv[1:]
    preview = "--preview" in args
    args    = [a for a in args if not a.startswith("--")]

    search_id = args[0] if args else None

    # 1. Locate the job folder
    job_dir = find_job_dir(search_id)
    print(f"  Job folder: {job_dir.name}")

    # 2. Find Excel and docx files (exclude Office lock files)
    xlsx_files = [f for f in job_dir.glob("*.xlsx") if not f.name.startswith("~$")]
    docx_files = [f for f in job_dir.glob("*.docx") if not f.name.startswith("~$")]

    if not xlsx_files:
        raise FileNotFoundError(f"No .xlsx file found in {job_dir}")
    if not docx_files:
        raise FileNotFoundError(f"No .docx file found in {job_dir}")

    excel_fn = xlsx_files[0]
    docx_fn  = docx_files[0]
    print(f"  Excel:      {excel_fn.name}")
    print(f"  Docx:       {docx_fn.name}")

    # 3. Reconstruct deal data from Excel
    wb             = load_workbook(excel_fn, data_only=True)
    folder_meta    = parse_folder_meta(job_dir)
    search_meta    = load_search_meta(wb, folder_meta)
    comp_summary   = load_comp_summary(wb)
    comm_spaces    = load_commercial_spaces(wb)

    print(f"\n  Search ID:     {search_meta['searchId']}")
    print(f"  Address:       {search_meta['address'][:70]}")
    print(f"  Email:         {search_meta['email']}")
    print(f"  Price:         ${float(search_meta['price']):,.0f}")
    print(f"  Units:         {search_meta['totalUnits']}")
    print(f"  Combos:        {len(comp_summary)}")
    print(f"  Commercial:    {len(comm_spaces)} space(s)")

    # 4. Generate email
    email_body = generate_email_body(
        search_meta, comp_summary, excel_fn, docx_fn,
        commercial_spaces=comm_spaces,
    )
    short_name = shorten_address(search_meta["address"])
    subject    = f"Your Ping analysis is ready \u2014 {short_name}"

    print(f"\n{'─' * 60}")
    print("  EMAIL BODY PREVIEW:")
    print(f"{'─' * 60}")
    print(email_body)
    print(f"{'─' * 60}\n")

    if preview:
        print("  [Preview mode] Email NOT sent. Run without --preview to send.")
        return

    print(f"  Sending to {search_meta['email']}...")
    send_email(search_meta["email"], subject, email_body, [excel_fn, docx_fn])
    print("  Done.")


if __name__ == "__main__":
    main()
