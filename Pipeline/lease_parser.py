"""
lease_parser.py — Parse lease documents and extract structured data.

Accepts file bytes + filename, detects file type, extracts text, and sends to OpenAI
to extract 24 lease-related fields as JSON.

Text-extraction strategy for PDFs (in order):
  1. pymupdf (fitz) — fast native text extraction; handles most modern PDFs.
  2. pytesseract OCR — for scanned / image-only PDFs where (1) yields < 100 chars.
"""

import io
import json
import logging
import os
from datetime import datetime

from openai import OpenAI

log = logging.getLogger("ping-server")


# Matches the original parse-lease.js prompt exactly
LEASE_FIELDS = [
    "tenant_name",
    "tenant_entity_type",
    "property_address",
    "unit_number",
    "asset_class",
    "lease_start_date",
    "lease_end_date",
    "lease_term_months",
    "base_rent_monthly",
    "base_rent_annual",
    "rent_escalation_type",
    "rent_escalation_value",
    "free_rent_months",
    "security_deposit",
    "expense_structure",
    "tenant_responsible_expenses",
    "landlord_responsible_expenses",
    "tenant_improvement_allowance",
    "renewal_options",
    "termination_option",
    "termination_notice_months",
    "guarantor_name",
    "commencement_conditions",
    "notes",
]

LEASE_EXTRACTION_PROMPT = (
    "You are a commercial real estate lease analyst. "
    "Extract the following fields from the provided lease document text. "
    'Return a JSON object with these exact field names as keys. If a field is not found in the text, use "Not found" as the value. '
    'Include a "confidence" field with an array of field names you are highly confident about. '
    'The "notes" field should contain any important caveats or ambiguities.\n\n'
    "Fields to extract:\n"
    + "\n".join(f"- {f}" for f in LEASE_FIELDS)
)

# ---------------------------------------------------------------------------
# Minimum characters threshold — if native extraction yields fewer chars
# than this, fall back to OCR (the PDF is likely scanned / image-only).
# ---------------------------------------------------------------------------
MIN_TEXT_CHARS = 100


# ---------------------------------------------------------------------------
# PDF text extraction
# ---------------------------------------------------------------------------

def _extract_text_from_pdf(file_bytes: bytes) -> str:
    """
    Extract text from a PDF.

    Strategy:
      1. Try pymupdf (fitz) — fast, handles most PDFs well.
      2. If that yields < MIN_TEXT_CHARS, fall back to OCR via pytesseract.
    """
    text = _extract_text_pymupdf(file_bytes)

    if len(text.strip()) >= MIN_TEXT_CHARS:
        return text.strip()

    # Native extraction got very little — likely a scanned PDF.  Try OCR.
    log.info("Native PDF text extraction yielded only %d chars — falling back to OCR", len(text.strip()))
    ocr_text = _extract_text_ocr(file_bytes)
    if ocr_text and len(ocr_text.strip()) > len(text.strip()):
        return ocr_text.strip()

    # Return whatever we got (even if short).
    return text.strip()


def _extract_text_pymupdf(file_bytes: bytes) -> str:
    """Extract text using pymupdf (fitz)."""
    try:
        import fitz  # pymupdf

        pdf_doc = fitz.open(stream=file_bytes, filetype="pdf")
        text = ""
        for page in pdf_doc:
            text += page.get_text() + "\n"
        pdf_doc.close()
        return text
    except Exception as e:
        log.error(f"pymupdf text extraction failed: {e}")
        # Fall through — caller will try OCR
        return ""


def _extract_text_ocr(file_bytes: bytes) -> str:
    """
    OCR a PDF using pdf2image + pytesseract.

    Requires system packages: tesseract-ocr, poppler-utils
    """
    try:
        from pdf2image import convert_from_bytes
        import pytesseract

        log.info("Running OCR on PDF (%d bytes)...", len(file_bytes))
        images = convert_from_bytes(file_bytes, dpi=300)
        text = ""
        for i, img in enumerate(images):
            page_text = pytesseract.image_to_string(img)
            text += page_text + "\n"
            log.info("OCR page %d: extracted %d chars", i + 1, len(page_text))
        return text
    except ImportError as e:
        log.warning("OCR dependencies not available (%s) — skipping OCR", e)
        return ""
    except Exception as e:
        log.error(f"OCR extraction failed: {e}")
        return ""


# ---------------------------------------------------------------------------
# Other file-type extractors (unchanged)
# ---------------------------------------------------------------------------

def _extract_text_from_csv(file_bytes: bytes) -> str:
    """Extract text from a CSV file."""
    try:
        text = file_bytes.decode("utf-8")
        return text.strip()
    except Exception as e:
        log.error(f"Error extracting CSV text: {e}")
        raise


def _extract_text_from_xlsx(file_bytes: bytes) -> str:
    """Extract text from an Excel file using openpyxl."""
    try:
        from openpyxl import load_workbook

        excel_file = io.BytesIO(file_bytes)
        workbook = load_workbook(excel_file)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"\n=== Sheet: {sheet_name} ===\n"
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        text += str(cell) + "\t"
                text += "\n"
        return text.strip()
    except Exception as e:
        log.error(f"Error extracting XLSX text: {e}")
        raise


def _extract_text_from_txt(file_bytes: bytes) -> str:
    """Extract text from a plain text file."""
    try:
        text = file_bytes.decode("utf-8")
        return text.strip()
    except Exception as e:
        log.error(f"Error extracting TXT text: {e}")
        raise


# ---------------------------------------------------------------------------
# Router
# ---------------------------------------------------------------------------

def _extract_text(file_bytes: bytes, filename: str) -> str:
    """Detect file type and extract text accordingly."""
    filename_lower = filename.lower()

    if filename_lower.endswith(".pdf"):
        return _extract_text_from_pdf(file_bytes)
    elif filename_lower.endswith(".csv"):
        return _extract_text_from_csv(file_bytes)
    elif filename_lower.endswith(".xlsx"):
        return _extract_text_from_xlsx(file_bytes)
    elif filename_lower.endswith(".txt"):
        return _extract_text_from_txt(file_bytes)
    else:
        raise ValueError(f"Unsupported file type: {filename}")


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def parse_lease(file_bytes: bytes, filename: str) -> dict:
    """
    Parse a lease document and extract 24 key fields using OpenAI.

    Args:
        file_bytes: The file contents as bytes
        filename: The filename (used to detect file type)

    Returns:
        {
            "parsed": {24-field lease object},
            "usage": {
                "prompt_tokens": int,
                "completion_tokens": int,
                "total_tokens": int,
                "estimated_cost": float
            }
        }
    """
    # Extract text from file
    log.info(f"Extracting text from {filename}...")
    extracted_text = _extract_text(file_bytes, filename)
    log.info(f"Extracted {len(extracted_text)} characters from {filename}")

    if len(extracted_text.strip()) < 20:
        raise ValueError(
            f"Could not extract meaningful text from {filename}. "
            "The file may be empty, corrupted, or in an unsupported format."
        )

    # Initialize OpenAI client
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise ValueError("OPENAI_API_KEY environment variable not set")

    client = OpenAI(api_key=api_key)

    # Call OpenAI with JSON mode
    log.info("Sending extracted text to OpenAI gpt-4o-mini...")
    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {
                "role": "system",
                "content": LEASE_EXTRACTION_PROMPT
            },
            {
                "role": "user",
                "content": extracted_text
            }
        ],
        response_format={"type": "json_object"},
        temperature=0
    )

    # Parse the response
    response_text = response.choices[0].message.content
    parsed_data = json.loads(response_text)

    # gpt-4o-mini pricing: $0.15 / 1M input, $0.60 / 1M output
    prompt_cost = (response.usage.prompt_tokens / 1_000_000) * 0.15
    completion_cost = (response.usage.completion_tokens / 1_000_000) * 0.60
    total_cost = prompt_cost + completion_cost

    log.info(f"Lease parsing complete. Tokens: {response.usage.total_tokens}, Cost: ${total_cost:.6f}")

    return {
        "parsed": parsed_data,
        "usage": {
            "prompt_tokens": response.usage.prompt_tokens,
            "completion_tokens": response.usage.completion_tokens,
            "total_tokens": response.usage.total_tokens,
            "estimated_cost": round(total_cost, 6)
        }
    }
